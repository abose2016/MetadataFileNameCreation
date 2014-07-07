__author__ = 'Anoushka'

from openpyxl import load_workbook, cell
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, Font
from openpyxl.cell import Cell, coordinate_from_string, column_index_from_string

#open workbook and get sheet
wb = load_workbook('Shiqmim_Photolog_Final.xlsx')
ws = wb.get_active_sheet()
numCols = 12 # Number of fields of metadata

for row in range(2, 1266): # loops through all rows of data
    name = "L:\Shiqmim Scanned Slides\\" + "19"
    yearRow = ws.cell('A%s'%row).value
    if(yearRow is not None):
        year = yearRow[len(yearRow) - 2: ] # saves year of each row of data
    else:
        year = ""
    for col_idx in range(1, numCols + 1): # loops through all fields for one particular row
        col = get_column_letter(col_idx)
        curr = ws.cell('%s%s'%(col, row)).value # gets current value of cell
        #year = curr[len(curr) - 2: ]
        if(curr is not None):
            if(col == 'A'):
                name += year + "\SHQ_" # adds year
            if(col == 'B'):
                name += year + "_" + curr[3:] # adds notebook location
            if(col == 'E'):
                name += "_" + str(curr) # adds area
            if(col == 'F'):
                name += "_ " + str(curr) # adds square
            if(col == 'G'):
                name += "L_" + str(curr) # adds locus
    name += ".tif"
    ws.cell('L%s'%(row)).value = name # saves file name to cell in column L

ws.cell('L1').value = "File Name" # Naming column
header_style = Style(font=Font(bold=True))
ws['L1'].style = header_style # Bolding text in column
wb.save('Shiqmim_Photolog_Final.xlsx')




